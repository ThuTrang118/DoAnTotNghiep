# utils/excel_utils.py
import os
import json
import csv
import sqlite3
import pandas as pd
from typing import List, Dict, Any, DefaultDict, Optional
from collections import defaultdict
from openpyxl import load_workbook

# NEW
import yaml
import xml.etree.ElementTree as ET

# ========== HÀM TIỆN ÍCH CƠ BẢN ==========
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def _norm(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ["nan", "none", "null"]:
        return ""
    return s

def _norm_key(s) -> str:
    return _norm(s).lower()

# ========== ĐỌC EXCEL (XLSX) ==========
def load_sheet(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_norm_key(h) for h in rows[0]]
    data = []
    for r in rows[1:]:
        if not r:
            continue
        row = {header[i]: _norm(r[i]) for i in range(min(len(header), len(r)))}
        if row.get("testcase"):
            data.append(row)

    print(f"[INFO] Loaded {len(data)} dòng từ sheet '{sheet_name}' ({os.path.basename(path)})")
    return data

# ========== ĐỌC XLS (Excel 97-2003) ==========
def load_xls(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    XLS cần xlrd. Cài:
      pip install xlrd==2.0.1
    """
    # sheet_name=None => đọc sheet đầu
    df = pd.read_excel(path, sheet_name=sheet_name or 0, engine="xlrd")
    df = df.where(pd.notnull(df), None)
    records = df.to_dict(orient="records")

    # chuẩn hoá key lower
    out = []
    for row in records:
        clean = {str(k).strip().lower(): _norm(v) for k, v in row.items()}
        if clean.get("testcase"):
            out.append(clean)

    print(f"[INFO] Loaded {len(out)} dòng từ XLS: {os.path.basename(path)}")
    return out

# ========== ĐỌC CSV ==========
def load_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data = []
        for row in reader:
            clean = {k.strip().lower(): _norm(v) for k, v in row.items()}
            if clean.get("testcase"):
                data.append(clean)
    print(f"[INFO] Loaded {len(data)} dòng từ CSV: {os.path.basename(path)}")
    return data

# ========== ĐỌC JSON ==========
def load_json(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # chấp nhận list hoặc {"items":[...]}
    if isinstance(raw, dict) and "items" in raw:
        raw = raw["items"]

    if not isinstance(raw, list):
        raise ValueError(f"JSON format not supported: {path}")

    data = []
    for item in raw:
        clean = {k.strip().lower(): _norm(v) for k, v in item.items()}
        if clean.get("testcase"):
            data.append(clean)
    print(f"[INFO] Loaded {len(data)} dòng từ JSON: {os.path.basename(path)}")
    return data

# ========== ĐỌC YAML/YML ==========
def load_yaml(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    # chấp nhận list hoặc {"items":[...]}
    if isinstance(raw, dict) and "items" in raw:
        raw = raw["items"]

    if not isinstance(raw, list):
        raise ValueError(f"YAML format not supported: {path}")

    data = []
    for item in raw:
        clean = {str(k).strip().lower(): _norm(v) for k, v in item.items()}
        if clean.get("testcase"):
            data.append(clean)

    print(f"[INFO] Loaded {len(data)} dòng từ YAML: {os.path.basename(path)}")
    return data

# ========== ĐỌC XML ==========
def load_xml(path: str, item_tag: str = "item") -> List[Dict[str, Any]]:
    """
    XML format đề xuất:
    <items>
      <item>
        <testcase>TC01</testcase>
        <username>...</username>
        <password>...</password>
        <expected>...</expected>
      </item>
    </items>
    """
    tree = ET.parse(path)
    root = tree.getroot()

    data = []
    for item in root.findall(f".//{item_tag}"):
        row = {}
        for child in list(item):
            row[str(child.tag).strip().lower()] = _norm(child.text)
        if row.get("testcase"):
            data.append(row)

    print(f"[INFO] Loaded {len(data)} dòng từ XML: {os.path.basename(path)}")
    return data

# ========== ĐỌC SQLITE/DB ==========
def load_sqlite(path: str, table: str = "testdata") -> List[Dict[str, Any]]:
    """
    DB (SQLite) khuyến nghị:
      - file: LoginData.db (hoặc .sqlite)
      - table: testdata
      - columns: testcase, username, password, expected, ...
    """
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM {table}")
        rows = cur.fetchall()
        data = []
        for r in rows:
            clean = {str(k).strip().lower(): _norm(r[k]) for k in r.keys()}
            if clean.get("testcase"):
                data.append(clean)
        print(f"[INFO] Loaded {len(data)} dòng từ DB table '{table}': {os.path.basename(path)}")
        return data
    finally:
        conn.close()

# ========== HÀM CHUNG ==========
def load_data(
    path: str,
    sheet_name: str = None,
    *,
    xml_item_tag: str = "item",
    db_table: str = "testdata",
) -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        return load_sheet(path, sheet_name or "Sheet1")

    if ext == ".xls":
        return load_xls(path, sheet_name=sheet_name)

    if ext == ".csv":
        return load_csv(path)

    if ext == ".json":
        return load_json(path)

    if ext in [".yaml", ".yml"]:
        return load_yaml(path)

    if ext == ".xml":
        return load_xml(path, item_tag=xml_item_tag)

    if ext in [".db", ".sqlite", ".sqlite3"]:
        return load_sqlite(path, table=db_table)

    raise ValueError(f"Không hỗ trợ định dạng: {ext}")

# ========== GHI KẾT QUẢ ==========
class ResultBook:
    def __init__(self, out_dir: str, file_name: str = "ResultsData.xlsx"):
        ensure_dir(out_dir)
        self.path = os.path.join(out_dir, file_name)
        self._sheets: DefaultDict[str, list[dict]] = defaultdict(list)

    def add_row(self, sheet: str, row: Dict[str, Any]):
        self._sheets[sheet].append(row)

    def save(self):
        existing = {}
        if os.path.exists(self.path):
            try:
                xls = pd.ExcelFile(self.path, engine="openpyxl")
                for s in xls.sheet_names:
                    existing[s] = xls.parse(s)
            except Exception:
                pass

        with pd.ExcelWriter(self.path, engine="openpyxl", mode="w") as writer:
            for s, df_old in existing.items():
                if s not in self._sheets:
                    df_old.to_excel(writer, sheet_name=s, index=False)
            for s, rows in self._sheets.items():
                pd.DataFrame(rows).to_excel(writer, sheet_name=s, index=False)

        print(f"[RESULT SAVED]: {self.path}")
        return self.path
