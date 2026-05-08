from __future__ import annotations

import csv
import json
import re
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from testdata_generation.engine.feature_item_schema import (
    build_default_testcase_id,
    get_feature_column_order,
    get_feature_item_fields,
    get_feature_output_basename,
    normalize_feature_name,
)

# =============================================================================
# SHARED STYLES FOR STEP1 EXCEL EXPORT
# =============================================================================
TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="C99700")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER_ALL = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

# =============================================================================
# STEP 1 COLUMN DEFINITIONS
# =============================================================================
STEP1_INTERNAL_COLUMNS = [
    "Coverage No",
    "Coverage ID",
    "Description",
    "Representative Value",
    "Expected Class",
    "Technique",
    "Rule",
    "Validity",
    "Partition Type",
    "Boundary Kind",
    "Boundary Reference",
    "Boundary Point",
]

STEP1_DISPLAY_HEADERS = [
    "No.",
    "Coverage ID",
    "Test Condition Description",
    "Test Data (Representative Value)",
    "Expected Result / Expected Behavior",
    "Testing Technique (EP/BVA)",
    "Validation Rule / Business Rule",
    "Data Validity (Valid/Invalid)",
    "Equivalence Partition Type",
    "Boundary Type",
    "Boundary Reference Value",
    "Boundary Point (MIN, MAX, N...)",
]


# =============================================================================
# STEP 3 / PROCESSED DATA EXPORTER
# =============================================================================
class DataExporter:
    """
    Exporter cho pipeline 3 bước.

    Mục tiêu:
    1. Raw/run artifacts:
       - lưu theo từng lần chạy trong output/<feature>_<timestamp>/
       - ví dụ:
         output/register_2026-04-19_10-15-32/
            step1.json
            step1.xlsx
            step2_dt.json
            step2_dt_invalid.json
            final.json
            final_invalid.json
            ...

    2. Processed data:
       - vẫn lưu vào data/ai_processed/<feature>/
       - vẫn giữ nguyên cơ chế chuyển đổi định dạng cho framework:
         json, csv, xlsx, xls, xml, yaml, yml, db
    """

    SUPPORTED_FORMATS = {"csv", "json", "xlsx", "xls", "xml", "yaml", "yml", "db"}

    def __init__(
        self,
        run_dir: Optional[str | Path] = None,
        processed_dir: Optional[str | Path] = None,
    ) -> None:
        project_root = Path(__file__).resolve().parents[2]

        self.run_dir = (
            Path(run_dir).resolve()
            if run_dir is not None
            else (project_root / "output")
        )

        self.processed_dir = (
            Path(processed_dir).resolve()
            if processed_dir is not None
            else (project_root / "data" / "ai_processed")
        )

        self.run_dir.mkdir(parents=True, exist_ok=True)
        self.processed_dir.mkdir(parents=True, exist_ok=True)

    # =========================================================
    # Helpers
    # =========================================================
    def _normalize_feature(self, feature: str) -> str:
        return normalize_feature_name(feature)

    def _processed_basename(self, feature: str) -> str:
        return get_feature_output_basename(feature)

    def _get_feature_processed_dir(self, feature: str) -> Path:
        feature_name = self._normalize_feature(feature)
        feature_dir = self.processed_dir / feature_name
        feature_dir.mkdir(parents=True, exist_ok=True)
        return feature_dir

    def _get_processed_json_path(self, feature: str) -> Path:
        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        return feature_dir / f"{name}.json"

    def _get_run_file_path(self, filename: str) -> Path:
        return self.run_dir / filename

    def _convert_final_testcase_to_framework_row(
        self,
        feature: str,
        item: Dict[str, Any],
        index: int,
    ) -> Dict[str, Any]:
        """
        Nhận cả 2 dạng:
        1. Schema mới Step 3:
           {
             "Testcase": "...",
             "<Field>": "...",
             "Expected": "..."
           }

        2. Schema cũ nếu còn sót:
           {
             "id": "...",
             "inputs": {...},
             "expected": "..."
           }

        Output luôn là row phẳng theo framework:
        Testcase + input fields + Expected
        """
        if not isinstance(item, dict):
            raise ValueError(f"Each exported item must be a dict, got: {type(item).__name__}")

        feature_fields = get_feature_item_fields(feature)

        testcase_id = item.get("Testcase")
        if not isinstance(testcase_id, str) or not testcase_id.strip():
            testcase_id = item.get("id")

        if not isinstance(testcase_id, str) or not testcase_id.strip():
            testcase_id = build_default_testcase_id(feature, index)

        inputs = item.get("inputs")
        if not isinstance(inputs, dict):
            inputs = item

        row: Dict[str, Any] = {"Testcase": testcase_id.strip()}

        for field in feature_fields:
            value = inputs.get(field, item.get(field, ""))
            row[field] = "" if value is None else value

        expected = item.get("Expected", item.get("expected", ""))
        row["Expected"] = "" if expected is None else str(expected)

        return row

    def _prepare_rows_for_processed(
        self,
        feature: str,
        rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        cleaned: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows, start=1):
            cleaned.append(self._convert_final_testcase_to_framework_row(feature, row, idx))
        return cleaned

    def _get_headers(self, feature: str) -> List[str]:
        return get_feature_column_order(feature)

    def _normalize_rows(
        self,
        rows: List[Dict[str, Any]],
        headers: List[str],
    ) -> List[Dict[str, Any]]:
        normalized: List[Dict[str, Any]] = []
        for row in rows:
            normalized_row: Dict[str, Any] = {}
            for h in headers:
                value = row.get(h, "")
                normalized_row[h] = "" if value is None else value
            normalized.append(normalized_row)
        return normalized

    def _load_processed_json_rows(self, feature: str) -> List[Dict[str, Any]]:
        path = self._get_processed_json_path(feature)
        if not path.exists():
            raise FileNotFoundError(f"Processed JSON not found: {path}")

        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            raise ValueError(f"Processed JSON must be an object: {path}")

        items = data.get("items", [])
        if not isinstance(items, list):
            raise ValueError(f"Processed JSON 'items' must be a list: {path}")

        rows: List[Dict[str, Any]] = []
        for idx, item in enumerate(items):
            if not isinstance(item, dict):
                raise ValueError(f"Processed JSON items[{idx}] must be an object: {path}")
            rows.append(item)

        return rows

    # =========================================================
    # RAW / RUN ARTIFACTS
    # =========================================================
    def write_raw_json(self, data: Dict[str, Any], filename: str) -> Path:
        """
        Ghi JSON raw/validated vào thư mục run hiện tại.
        Ví dụ:
        - step1.json
        - step1_invalid.json
        - step2_dt.json
        - step2_dt_invalid.json
        - final.json
        - final_invalid.json
        """
        path = self._get_run_file_path(filename)
        with path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return path

    # =========================================================
    # PROCESSED JSON (SOURCE OF TRUTH FOR CONVERSIONS)
    # =========================================================
    def write_processed_json(self, feature: str, rows: List[Dict[str, Any]]) -> Path:
        """
        Ghi processed JSON vào data/ai_processed/<feature>/...
        Đây là source of truth cho các chuyển đổi định dạng còn lại.
        """
        path = self._get_processed_json_path(feature)

        rows = self._prepare_rows_for_processed(feature, rows)
        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        with path.open("w", encoding="utf-8") as f:
            json.dump({"items": rows_norm}, f, ensure_ascii=False, indent=2)

        return path

    # =========================================================
    # TỪ PROCESSED JSON -> CÁC ĐỊNH DẠNG KHÁC (GIỮ NGUYÊN LOGIC)
    # =========================================================
    def write_processed_csv_from_json(self, feature: str) -> Path:
        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.csv"

        rows = self._load_processed_json_rows(feature)
        if not rows:
            path.write_text("", encoding="utf-8")
            return path

        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows_norm)

        return path

    def write_processed_xlsx_from_json(self, feature: str) -> Path:
        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.xlsx"

        rows = self._load_processed_json_rows(feature)
        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        df = pd.DataFrame(rows_norm, columns=headers)
        sheet_name = name[:31] if name else "Sheet1"
        df.to_excel(path, index=False, sheet_name=sheet_name)

        return path

    def write_processed_xls_from_json(self, feature: str) -> Path:
        try:
            import xlwt  # type: ignore
        except Exception as e:
            raise RuntimeError(
                "Cannot write .xls. Install in venv: pip install xlwt. "
                f"Original error: {e}"
            )

        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.xls"

        rows = self._load_processed_json_rows(feature)
        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        wb = xlwt.Workbook()
        ws = wb.add_sheet(name[:31] if name else "Sheet1")

        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header)

        for row_idx, row in enumerate(rows_norm, start=1):
            for col_idx, header in enumerate(headers):
                value = row.get(header, "")
                ws.write(row_idx, col_idx, "" if value is None else str(value))

        wb.save(str(path))
        return path

    def write_processed_xml_from_json(self, feature: str) -> Path:
        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.xml"

        rows = self._load_processed_json_rows(feature)
        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        def esc(s: str) -> str:
            return (
                s.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace('"', "&quot;")
                .replace("'", "&apos;")
            )

        lines = ["<items>"]
        for row in rows_norm:
            lines.append("  <item>")
            for key in headers:
                value = row.get(key, "")
                lines.append(f"    <{key}>{esc('' if value is None else str(value))}</{key}>")
            lines.append("  </item>")
        lines.append("</items>")

        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    def write_processed_yaml_from_json(self, feature: str, ext: str) -> Path:
        try:
            import yaml  # type: ignore
        except Exception as e:
            raise RuntimeError(
                f"Cannot write YAML. Install: pip install pyyaml. Original error: {e}"
            )

        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.{ext}"

        rows = self._load_processed_json_rows(feature)
        headers = self._get_headers(feature)
        rows_norm = self._normalize_rows(rows, headers)

        with path.open("w", encoding="utf-8") as f:
            yaml.safe_dump({"items": rows_norm}, f, allow_unicode=True, sort_keys=False)

        return path

    def write_processed_db_from_json(self, feature: str) -> Path:
        name = self._processed_basename(feature)
        feature_dir = self._get_feature_processed_dir(feature)
        path = feature_dir / f"{name}.db"
        table = self._normalize_feature(feature) or "testdata"

        rows = self._load_processed_json_rows(feature)

        conn = sqlite3.connect(path)
        try:
            cur = conn.cursor()
            cur.execute(f'DROP TABLE IF EXISTS "{table}"')

            if not rows:
                conn.commit()
                return path

            cols = self._get_headers(feature)
            rows_norm = self._normalize_rows(rows, cols)

            col_defs = ", ".join([f'"{c}" TEXT' for c in cols])
            cur.execute(f'CREATE TABLE "{table}" ({col_defs})')

            placeholders = ", ".join(["?"] * len(cols))
            col_names = ", ".join([f'"{c}"' for c in cols])
            sql = f'INSERT INTO "{table}" ({col_names}) VALUES ({placeholders})'

            values = []
            for row in rows_norm:
                values.append(["" if row.get(c) is None else str(row.get(c)) for c in cols])

            cur.executemany(sql, values)
            conn.commit()
            return path
        finally:
            conn.close()

    # =========================================================
    # PUBLIC EXPORT API
    # =========================================================
    def write_formats(
        self,
        feature: str,
        rows: List[Dict[str, Any]],
        formats: Optional[Iterable[str]] = None,
        yaml_ext: str = "yaml",
    ) -> Dict[str, Path]:
        """
        Logic giữ nguyên:
        1. ghi processed JSON vào data/ai_processed/<feature>/
        2. từ processed JSON, convert sang các định dạng khác trong cùng thư mục đó
        """
        if formats is None:
            formats = sorted(self.SUPPORTED_FORMATS)

        normalized_formats = []
        for fmt in formats:
            f = (fmt or "").strip().lower()
            if f in self.SUPPORTED_FORMATS and f not in normalized_formats:
                normalized_formats.append(f)

        out: Dict[str, Path] = {}

        json_path = self.write_processed_json(feature, rows)
        out["json"] = json_path

        for f in normalized_formats:
            if f == "json":
                continue
            if f == "csv":
                out["csv"] = self.write_processed_csv_from_json(feature)
            elif f == "xlsx":
                out["xlsx"] = self.write_processed_xlsx_from_json(feature)
            elif f == "xls":
                out["xls"] = self.write_processed_xls_from_json(feature)
            elif f == "xml":
                out["xml"] = self.write_processed_xml_from_json(feature)
            elif f == "yaml":
                out["yaml"] = self.write_processed_yaml_from_json(feature, ext=yaml_ext)
            elif f == "yml":
                out["yml"] = self.write_processed_yaml_from_json(feature, ext="yml")
            elif f == "db":
                out["db"] = self.write_processed_db_from_json(feature)

        return out

    def export(
        self,
        feature: str,
        rows: List[Dict[str, Any]],
        formats: Optional[Iterable[str]] = None,
    ) -> Dict[str, str]:
        paths = self.write_formats(feature=feature, rows=rows, formats=formats)
        return {k: str(v) for k, v in paths.items()}

    def export_feature_items(
        self,
        feature: str,
        items: List[Dict[str, Any]],
        formats: Optional[Iterable[str]] = None,
    ) -> List[str]:
        paths = self.write_formats(feature=feature, rows=items, formats=formats)
        return [str(v) for v in paths.values()]


class DataWriter(DataExporter):
    pass


# =============================================================================
# STEP 1 COVERAGE -> EXCEL EXPORTER
# =============================================================================
class Step1ExcelExporter:
    """
    Export Step 1 coverage JSON -> multi-sheet Excel.

    Giữ 2 mục tiêu:
    1. Sheet kỹ thuật để debug pipeline
    2. Sheet trình bày gần giống mẫu bài giảng EP + BVA:
       Test case | Field | EP | BVA | Test Data | Expected Class
    """

    def load_step1_json(self, json_path: str | Path) -> Dict[str, Any]:
        path = Path(json_path)
        if not path.exists():
            raise FileNotFoundError(f"JSON file not found: {path}")

        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            raise ValueError("Step 1 JSON must be an object")
        if "feature" not in data:
            raise ValueError("Step 1 JSON must contain 'feature'")
        if "coverage_items" not in data or not isinstance(data["coverage_items"], list):
            raise ValueError("Step 1 JSON must contain 'coverage_items' as a list")

        return data

    @staticmethod
    def as_clean_str(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def normalize_sheet_name(name: str, used_names: set[str]) -> str:
        cleaned = INVALID_SHEET_CHARS.sub("_", name.strip() or "Sheet")
        cleaned = cleaned[:31] or "Sheet"

        candidate = cleaned
        counter = 2
        while candidate in used_names:
            suffix = f"_{counter}"
            candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
            counter += 1

        used_names.add(candidate)
        return candidate

    @staticmethod
    def sort_boundary_point(point: str) -> int:
        ranking = {
            "MIN-1": 1,
            "MIN": 2,
            "MIN+1": 3,
            "MAX-1": 4,
            "MAX": 5,
            "MAX+1": 6,
            "N-1": 7,
            "N": 8,
            "N+1": 9,
        }
        return ranking.get((point or "").upper(), 99)

    def build_field_rows(self, data: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        counters: Dict[str, int] = defaultdict(int)

        coverage_items = data.get("coverage_items", [])
        for item in coverage_items:
            if not isinstance(item, dict):
                continue

            field = self.as_clean_str(item.get("field")) or "General"
            counters[field] += 1
            boundary = item.get("boundary") if isinstance(item.get("boundary"), dict) else {}

            grouped[field].append(
                {
                    "Coverage No": f"C{counters[field]}",
                    "Coverage ID": self.as_clean_str(item.get("id")),
                    "Description": self.as_clean_str(item.get("description")),
                    "Representative Value": self.as_clean_str(item.get("representative_value")),
                    "Expected Class": self.as_clean_str(item.get("expected_class")),
                    "Technique": self.as_clean_str(item.get("technique")).upper(),
                    "Rule": self.as_clean_str(item.get("rule")),
                    "Validity": self.as_clean_str(item.get("validity")),
                    "Partition Type": self.as_clean_str(item.get("partition_type")),
                    "Boundary Kind": self.as_clean_str(boundary.get("kind")),
                    "Boundary Reference": self.as_clean_str(boundary.get("reference")),
                    "Boundary Point": self.as_clean_str(boundary.get("point")).upper(),
                    "Field": field,
                }
            )

        for field, rows in grouped.items():
            rows.sort(
                key=lambda row: (
                    0 if row["Technique"] == "EP" else 1,
                    self.sort_boundary_point(row["Boundary Point"]),
                    row["Coverage No"],
                )
            )
            for idx, row in enumerate(rows, start=1):
                row["Coverage No"] = f"C{idx}"

        return dict(grouped)

    def build_lecture_rows(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []

        coverage_items = data.get("coverage_items", [])
        normalized_items: List[Dict[str, Any]] = []
        for item in coverage_items:
            if isinstance(item, dict):
                normalized_items.append(item)

        def lecture_sort_key(item: Dict[str, Any]) -> tuple:
            technique = self.as_clean_str(item.get("technique")).upper()
            field = self.as_clean_str(item.get("field"))
            boundary = item.get("boundary") if isinstance(item.get("boundary"), dict) else {}
            point = self.as_clean_str(boundary.get("point")).upper()
            rule = self.as_clean_str(item.get("rule"))
            return (
                field,
                0 if technique == "EP" else 1,
                self.sort_boundary_point(point),
                rule,
            )

        normalized_items.sort(key=lecture_sort_key)

        for idx, item in enumerate(normalized_items, start=1):
            technique = self.as_clean_str(item.get("technique")).upper()
            field = self.as_clean_str(item.get("field"))
            rule = self.as_clean_str(item.get("rule"))
            representative_value = self.as_clean_str(item.get("representative_value"))
            expected_class = self.as_clean_str(item.get("expected_class"))
            validity = self.as_clean_str(item.get("validity")).lower()
            boundary = item.get("boundary") if isinstance(item.get("boundary"), dict) else {}
            point = self.as_clean_str(boundary.get("point")).upper()

            ep_text = ""
            bva_text = ""

            if technique == "EP":
                ep_text = rule or self.as_clean_str(item.get("description"))
            elif technique == "BVA":
                bva_text = point
                if validity == "valid":
                    ep_text = "Hợp lệ"
                else:
                    ep_text = "Không hợp lệ"

            rows.append(
                {
                    "Test case": f"TC{idx}",
                    "Field": field,
                    "EP": ep_text,
                    "BVA": bva_text,
                    "Test Data": representative_value,
                    "Expected Class": expected_class,
                    "Coverage ID": self.as_clean_str(item.get("id")),
                    "Technique": technique,
                }
            )

        return rows

    @staticmethod
    def apply_table_header(cell) -> None:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER_ALL

    @staticmethod
    def apply_data_cell(cell, align=LEFT) -> None:
        cell.alignment = align
        cell.border = BORDER_ALL

    @staticmethod
    def auto_fit_columns(ws) -> None:
        widths: Dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                widths[cell.column] = max(widths.get(cell.column, 0), min(len(text) + 2, 45))

        for col_idx, width in widths.items():
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(width, 12)

    def create_summary_sheet(
        self,
        wb: Workbook,
        feature_name: str,
        data: Dict[str, Any],
        field_rows: Dict[str, List[Dict[str, Any]]],
    ) -> None:
        ws = wb.active
        ws.title = "Summary"

        summary = data.get("coverage_summary") if isinstance(data.get("coverage_summary"), dict) else {}
        ep_count = summary.get("EP_count", "")
        bva_count = summary.get("BVA_count", "")
        total_count = summary.get("TOTAL", len(data.get("coverage_items", [])))
        description = self.as_clean_str(data.get("description"))

        ws["A1"] = f"STEP 1 COVERAGE SUMMARY - {feature_name.upper()}"
        ws["A1"].fill = TITLE_FILL
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = LEFT
        ws.merge_cells("A1:F1")

        meta_rows = [
            ("Feature", feature_name),
            ("Description", description),
            ("EP Items", ep_count),
            ("BVA Items", bva_count),
            ("TOTAL", total_count),
            ("Số field", len(field_rows)),
        ]

        start_meta_row = 3
        for idx, (label, value) in enumerate(meta_rows, start=start_meta_row):
            ws.cell(idx, 1, label)
            ws.cell(idx, 2, value)
            ws.cell(idx, 1).fill = SECTION_FILL
            ws.cell(idx, 1).font = SUBHEADER_FONT
            ws.cell(idx, 1).alignment = LEFT
            ws.cell(idx, 2).alignment = WRAP_LEFT if label == "Description" else LEFT
            ws.cell(idx, 1).border = BORDER_ALL
            ws.cell(idx, 2).border = BORDER_ALL

        table_row = 11
        headers = ["Field", "Total Items", "EP Items", "BVA Items", "Sheet Name"]
        for col, header in enumerate(headers, start=1):
            self.apply_table_header(ws.cell(table_row, col, header))

        used_sheet_names = {"Summary"}
        row = table_row + 1

        for values in [
            ["Bảng trình bày", len(data.get("coverage_items", [])), ep_count, bva_count, "Bang_EP_BVA"],
        ]:
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                self.apply_data_cell(cell, LEFT if col in {1, 5} else CENTER)
            row += 1

        for field, rows in field_rows.items():
            ep_items = sum(1 for r in rows if r["Technique"] == "EP")
            bva_items = sum(1 for r in rows if r["Technique"] == "BVA")
            sheet_name = self.normalize_sheet_name(field, used_sheet_names)

            values = [
                field,
                len(rows),
                ep_items,
                bva_items,
                sheet_name,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                self.apply_data_cell(cell, LEFT if col in {1, 5} else CENTER)
            row += 1

        ws.freeze_panes = "A12"
        ws.auto_filter.ref = f"A11:E{max(row - 1, 11)}"
        self.auto_fit_columns(ws)

    def create_lecture_view_sheet(
        self,
        wb: Workbook,
        feature_name: str,
        data: Dict[str, Any],
        used_names: set[str],
    ) -> None:
        sheet_name = self.normalize_sheet_name("Bang_EP_BVA", used_names)
        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = f"BẢNG EP + BVA - {feature_name.upper()}"
        ws["A1"].fill = TITLE_FILL
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = LEFT
        ws.merge_cells("A1:H1")

        ws["A3"] = "Bảng này trình bày Step 1 theo dạng gần giống mẫu bài giảng."
        ws["A3"].alignment = LEFT
        ws.merge_cells("A3:H3")

        ws["A5"] = "Test case"
        ws["B5"] = "Field"
        ws["C5"] = "Techniques"
        ws["E5"] = "Test Data"
        ws["F5"] = "Expected Class"
        ws["G5"] = "Coverage ID"
        ws["H5"] = "Technique"

        ws.merge_cells("A5:A6")
        ws.merge_cells("B5:B6")
        ws.merge_cells("C5:D5")
        ws.merge_cells("E5:E6")
        ws.merge_cells("F5:F6")
        ws.merge_cells("G5:G6")
        ws.merge_cells("H5:H6")

        ws["C6"] = "EP"
        ws["D6"] = "BVA"

        for cell_ref in ["A5", "B5", "C5", "E5", "F5", "G5", "H5", "C6", "D6"]:
            self.apply_table_header(ws[cell_ref])

        rows = self.build_lecture_rows(data)
        start_row = 7
        for idx, row in enumerate(rows, start=start_row):
            values = [
                row["Test case"],
                row["Field"],
                row["EP"],
                row["BVA"],
                row["Test Data"],
                row["Expected Class"],
                row["Coverage ID"],
                row["Technique"],
            ]
            aligns = [CENTER, LEFT, WRAP_LEFT, CENTER, WRAP_LEFT, WRAP_LEFT, CENTER, CENTER]
            for col, (value, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(idx, col, value)
                self.apply_data_cell(cell, align)

        ws.freeze_panes = "A7"
        ws.auto_filter.ref = f"A6:H{max(start_row + len(rows) - 1, 6)}"

        widths = {
            "A": 12,
            "B": 18,
            "C": 28,
            "D": 14,
            "E": 20,
            "F": 28,
            "G": 14,
            "H": 12,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def create_field_sheet(
        self,
        wb: Workbook,
        feature_name: str,
        field_name: str,
        rows: List[Dict[str, Any]],
        used_names: set[str],
    ) -> None:
        sheet_name = self.normalize_sheet_name(field_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        title = f"STEP 1 - {feature_name.upper()} - {field_name}"
        ws["A1"] = title
        ws["A1"].fill = TITLE_FILL
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = LEFT
        ws.merge_cells("A1:L1")

        counts = Counter(row["Technique"] for row in rows)
        meta = [
            ("Feature", feature_name),
            ("Field", field_name),
            ("EP Items", counts.get("EP", 0)),
            ("BVA Items", counts.get("BVA", 0)),
            ("Total", len(rows)),
        ]

        meta_start_row = 3
        for idx, (label, value) in enumerate(meta, start=meta_start_row):
            ws.cell(idx, 1, label)
            ws.cell(idx, 2, value)
            ws.cell(idx, 1).fill = SECTION_FILL
            ws.cell(idx, 1).font = SUBHEADER_FONT
            ws.cell(idx, 1).border = BORDER_ALL
            ws.cell(idx, 2).border = BORDER_ALL
            ws.cell(idx, 1).alignment = LEFT
            ws.cell(idx, 2).alignment = LEFT if idx < meta_start_row + 2 else CENTER

        header_row = 9
        for col, header in enumerate(STEP1_DISPLAY_HEADERS, start=1):
            self.apply_table_header(ws.cell(header_row, col, header))

        current_row = header_row + 1
        for row in rows:
            values = [
                row["Coverage No"],
                row["Coverage ID"],
                row["Description"],
                row["Representative Value"],
                row["Expected Class"],
                row["Technique"],
                row["Rule"],
                row["Validity"],
                row["Partition Type"],
                row["Boundary Kind"],
                row["Boundary Reference"],
                row["Boundary Point"],
            ]
            aligns = [
                CENTER,
                CENTER,
                WRAP_LEFT,
                WRAP_LEFT,
                WRAP_LEFT,
                CENTER,
                WRAP_LEFT,
                CENTER,
                CENTER,
                CENTER,
                CENTER,
                CENTER,
            ]
            for col, (value, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(current_row, col, value)
                self.apply_data_cell(cell, align)
            current_row += 1

        ws.freeze_panes = "A10"
        ws.auto_filter.ref = f"A9:L{max(current_row - 1, 9)}"

        manual_widths = {
            "A": 10,
            "B": 14,
            "C": 38,
            "D": 24,
            "E": 30,
            "F": 18,
            "G": 32,
            "H": 20,
            "I": 24,
            "J": 16,
            "K": 22,
            "L": 24,
        }
        for col, width in manual_widths.items():
            ws.column_dimensions[col].width = width

    def export_step1_to_excel(self, json_path: str | Path, output_path: str | Path) -> Path:
        data = self.load_step1_json(json_path)
        feature_name = self.as_clean_str(data.get("feature")) or "feature"
        field_rows = self.build_field_rows(data)

        wb = Workbook()
        self.create_summary_sheet(wb, feature_name, data, field_rows)

        used_names = {"Summary"}
        self.create_lecture_view_sheet(wb, feature_name, data, used_names)

        for field_name, rows in field_rows.items():
            self.create_field_sheet(wb, feature_name, field_name, rows, used_names)

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return output


def export_step1_to_excel(json_path: str | Path, output_path: str | Path) -> Path:
    exporter = Step1ExcelExporter()
    return exporter.export_step1_to_excel(json_path, output_path)

# =============================================================================
# STEP 2 DECISION TABLE -> EXCEL EXPORTER
# =============================================================================
class Step2DecisionTableExcelExporter:
    """
    Export Step 2 Decision Table ra Excel.

    Mục tiêu sửa ở Step 2:
    - Không hiển thị mã kỹ thuật C1, C2... ở cột condition.
    - Không hiển thị mã kỹ thuật A1, A2... ở dòng action.
    - Không hiển thị R1, R2... ở header rule.
    - Thay bằng mô tả tiếng Việt cụ thể lấy trực tiếp từ step2_dt.json:
      conditions[].name, actions[].name/expected, decision_rules[].reduction_note/expected.

    Lưu ý:
    - Không động vào Step 1 exporter phía trên.
    - JSON Step 2 vẫn giữ id C1/A1/DT_001 để pipeline trace và validate.
      Việc thay mô tả chỉ áp dụng khi trình bày ra Excel.
    """

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="00B0F0")
    SECTION_FILL = PatternFill(fill_type="solid", fgColor="CCFFFF")
    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    BLACK_SIDE = Side(style="thin", color="000000")
    BORDER = Border(
        left=BLACK_SIDE,
        right=BLACK_SIDE,
        top=BLACK_SIDE,
        bottom=BLACK_SIDE,
    )

    def _clean(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _normalize_state(self, value: Any) -> str:
        raw = self._clean(value).upper()
        if raw in {"Y", "YES", "TRUE", "1"}:
            return "Y"
        if raw in {"N", "NO", "FALSE", "0"}:
            return "N"
        if raw == "-":
            return "-"
        return ""

    def _extract_conditions(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        out: List[Dict[str, str]] = []
        raw_conditions = data.get("conditions", [])
        if not isinstance(raw_conditions, list):
            return out

        for idx, cond in enumerate(raw_conditions, start=1):
            if not isinstance(cond, dict):
                continue

            condition_id = self._clean(cond.get("id")) or f"C{idx}"
            condition_name = self._clean(cond.get("name")) or "Điều kiện"
            values = cond.get("values")

            if isinstance(values, list) and values:
                value_text = "/".join(self._clean(v) for v in values if self._clean(v))
            else:
                value_text = "Y/N"

            out.append(
                {
                    "id": condition_id,
                    "label": f"{condition_id} - {condition_name}",
                    "values": value_text or "Y/N",
                }
            )

        return out

    def _extract_actions(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        out: List[Dict[str, str]] = []
        raw_actions = data.get("actions", [])
        if not isinstance(raw_actions, list):
            return out

        for idx, action in enumerate(raw_actions, start=1):
            if not isinstance(action, dict):
                continue

            action_id = self._clean(action.get("id")) or f"A{idx}"
            action_name = self._clean(action.get("name"))
            expected = self._clean(action.get("expected"))

            if expected and action_name:
                label = f"{action_id} - {action_name}: {expected}"
            elif expected:
                label = f"{action_id} - {expected}"
            elif action_name:
                label = f"{action_id} - {action_name}"
            else:
                label = action_id

            out.append({"id": action_id, "label": label})

        return out

    def _extract_rules(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        raw_rules = data.get("decision_rules", [])
        if not isinstance(raw_rules, list):
            return []
        return [rule for rule in raw_rules if isinstance(rule, dict)]

    def _get_condition_state(self, rule: Dict[str, Any], condition_id: str) -> str:
        states = rule.get("condition_states", {})
        if not isinstance(states, dict):
            return ""
        return self._normalize_state(states.get(condition_id))

    def _get_action_refs(self, rule: Dict[str, Any]) -> List[str]:
        refs = rule.get("action_refs", [])
        if not isinstance(refs, list):
            return []
        return [self._clean(ref) for ref in refs if self._clean(ref)]

    def _get_rule_label(self, rule: Dict[str, Any], index: int) -> str:
        rule_id = self._clean(rule.get("id")) or f"DT_{index:03d}"
        rule_type = self._clean(rule.get("type"))
        expected = self._clean(rule.get("expected"))
        note = self._clean(rule.get("reduction_note"))

        details = note or expected or rule_type

        if details:
            return f"{rule_id}\n{details}"

        return rule_id

    def _set_cell(self, ws, row: int, col: int, value: Any, *, fill=None, font=None, alignment=None) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = self.BORDER
        cell.fill = fill or self.WHITE_FILL
        cell.font = font or Font(size=11)
        cell.alignment = alignment or WRAP_CENTER

    def export_step2_to_excel(self, data: Dict[str, Any], output_path: str | Path) -> Path:
        if not isinstance(data, dict):
            raise ValueError("Step2 data must be a dict")

        conditions = self._extract_conditions(data)
        actions = self._extract_actions(data)
        rules = self._extract_rules(data)

        if not conditions:
            raise ValueError("conditions rỗng")
        if not actions:
            raise ValueError("actions rỗng")
        if not rules:
            raise ValueError("decision_rules rỗng")

        wb = Workbook()
        ws = wb.active
        ws.title = "Decision_Table"
        ws.sheet_view.showGridLines = False

        total_rules = len(rules)
        first_rule_col = 3
        last_col = first_rule_col + total_rules - 1

        # ===== TITLE =====
        title = self._clean(data.get("description")) or "BẢNG QUYẾT ĐỊNH STEP 2"
        self._set_cell(
            ws,
            1,
            1,
            title,
            fill=self.HEADER_FILL,
            font=Font(bold=True, size=14),
            alignment=LEFT,
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

        # ===== HEADER =====
        self._set_cell(
            ws,
            3,
            first_rule_col,
            "Các luật quyết định",
            fill=self.HEADER_FILL,
            font=Font(bold=True, size=12),
        )
        if last_col > first_rule_col:
            ws.merge_cells(start_row=3, start_column=first_rule_col, end_row=3, end_column=last_col)

        self._set_cell(ws, 4, 1, "Điều kiện / Nguyên nhân", fill=self.SECTION_FILL, font=Font(bold=True, size=12), alignment=WRAP_CENTER)
        self._set_cell(ws, 4, 2, "Giá trị", fill=self.SECTION_FILL, font=Font(bold=True, size=12), alignment=WRAP_CENTER)

        for idx, rule in enumerate(rules, start=1):
            self._set_cell(
                ws,
                4,
                first_rule_col + idx - 1,
                self._get_rule_label(rule, idx),
                fill=self.SECTION_FILL,
                font=Font(bold=True, size=10),
                alignment=WRAP_CENTER,
            )

        # ===== CONDITIONS =====
        row = 5
        for cond in conditions:
            self._set_cell(ws, row, 1, cond["label"], alignment=WRAP_LEFT)
            self._set_cell(ws, row, 2, cond["values"], alignment=CENTER)

            for col, rule in enumerate(rules, start=first_rule_col):
                self._set_cell(ws, row, col, self._get_condition_state(rule, cond["id"]), alignment=CENTER)
            row += 1

        # ===== ACTIONS =====
        self._set_cell(ws, row, 1, "Kết quả / Hành động", fill=self.SECTION_FILL, font=Font(bold=True, size=12), alignment=WRAP_CENTER)
        self._set_cell(ws, row, 2, "", fill=self.SECTION_FILL, font=Font(bold=True, size=12), alignment=WRAP_CENTER)
        for col in range(first_rule_col, last_col + 1):
            self._set_cell(ws, row, col, "", fill=self.SECTION_FILL, font=Font(bold=True, size=12), alignment=WRAP_CENTER)
        row += 1

        for action in actions:
            self._set_cell(ws, row, 1, action["label"], alignment=WRAP_LEFT)
            self._set_cell(ws, row, 2, "", alignment=CENTER)

            for col, rule in enumerate(rules, start=first_rule_col):
                mark = "X" if action["id"] in self._get_action_refs(rule) else ""
                self._set_cell(ws, row, col, mark, alignment=CENTER, font=Font(bold=True, size=11))
            row += 1

        # ===== LAYOUT =====
        ws.freeze_panes = "C5"
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 12
        for col in range(first_rule_col, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 28

        for row_idx in range(1, row + 1):
            ws.row_dimensions[row_idx].height = 34
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[4].height = 55

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return output


def export_step2_to_excel(data: Dict[str, Any], output_path: str | Path) -> Path:
    exporter = Step2DecisionTableExcelExporter()
    return exporter.export_step2_to_excel(data, output_path)
